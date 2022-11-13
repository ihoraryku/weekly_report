import shutil
from datetime import datetime, timedelta
import easygui
from openpyxl import load_workbook

easygui.msgbox('Вас вітає програма для автоматизованого створення щонедільного звіту для НЕК. Оберіть необхідну теку (місяць) де зберігаються файли НЕК')

PATH_CHOICE = easygui.diropenbox()

FIRST_DATE = (datetime.today() + timedelta(1)).strftime("%d.%m.%Y")
SECOND_DATE = (datetime.today() + timedelta(2)).strftime("%d.%m.%Y")
THIRD_DATE = (datetime.today() + timedelta(3)).strftime("%d.%m.%Y")
FOURTH_DATE = (datetime.today() + timedelta(4)).strftime("%d.%m.%Y")
FIFTH_DATE = (datetime.today() + timedelta(5)).strftime("%d.%m.%Y")
SIXTH_DATE = (datetime.today() + timedelta(6)).strftime("%d.%m.%Y")
SEVENTH_DATE = (datetime.today() + timedelta(7)).strftime("%d.%m.%Y")

DEST_DIR = rf'{PATH_CHOICE}\Недельный на {FIRST_DATE}-{SEVENTH_DATE}'
print(f'path_choice: {PATH_CHOICE}')

src_dir_templates = r"C:\Users\ihoraryku\Downloads\weekly_report\templates"
shutil.copytree(src_dir_templates, DEST_DIR)


def createBolgradSolar():

    scr_filename = 'Bolgrad_Solar.xlsx'
    dest_filename = 'Болград Солар.xlsx'

    # First date
    scr_file = rf'{PATH_CHOICE}\{FIRST_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=3, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # Second date
    scr_file = rf'{PATH_CHOICE}\{SECOND_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=4, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # third_date
    scr_file = rf'{PATH_CHOICE}\{THIRD_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=5, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # fourth_date
    scr_file = rf'{PATH_CHOICE}\{FOURTH_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=6, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # fifth_date
    scr_file = rf'{PATH_CHOICE}\{FIFTH_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=7, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # sixth_date
    scr_file = rf'{PATH_CHOICE}\{SIXTH_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=8, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # seventh_date
    scr_file = rf'{PATH_CHOICE}\{SEVENTH_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=9, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

createBolgradSolar()

def createVoshodSolar():

    scr_filename = 'Voshod_Solar.xlsx'
    dest_filename = 'Восход Солар.xlsx'

    # First date
    scr_file = rf'{PATH_CHOICE}\{FIRST_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=3, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # Second date
    scr_file = rf'{PATH_CHOICE}\{SECOND_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=4, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # third_date
    scr_file = rf'{PATH_CHOICE}\{THIRD_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=5, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # fourth_date
    scr_file = rf'{PATH_CHOICE}\{FOURTH_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=6, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # fifth_date
    scr_file = rf'{PATH_CHOICE}\{FIFTH_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=7, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # sixth_date
    scr_file = rf'{PATH_CHOICE}\{SIXTH_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=8, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # seventh_date
    scr_file = rf'{PATH_CHOICE}\{SEVENTH_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=9, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)
createVoshodSolar()

def createDNfirst():

    scr_filename = 'Dunaiskaya_SES1.xlsx'
    dest_filename = 'Дунайская СЭС-1.xlsx'

    # First date
    scr_file = rf'{PATH_CHOICE}\{FIRST_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=3, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # Second date
    scr_file = rf'{PATH_CHOICE}\{SECOND_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=4, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # third_date
    scr_file = rf'{PATH_CHOICE}\{THIRD_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=5, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # fourth_date
    scr_file = rf'{PATH_CHOICE}\{FOURTH_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=6, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # fifth_date
    scr_file = rf'{PATH_CHOICE}\{FIFTH_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=7, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # sixth_date
    scr_file = rf'{PATH_CHOICE}\{SIXTH_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=8, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # seventh_date
    scr_file = rf'{PATH_CHOICE}\{SEVENTH_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=9, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)
createDNfirst()

def createDNsecond():

    scr_filename = 'Dunaiskaya_SES2.xlsx'
    dest_filename = 'Дунайская СЭС-2.xlsx'

    # First date
    scr_file = rf'{PATH_CHOICE}\{FIRST_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=3, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # Second date
    scr_file = rf'{PATH_CHOICE}\{SECOND_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=4, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # third_date
    scr_file = rf'{PATH_CHOICE}\{THIRD_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=5, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # fourth_date
    scr_file = rf'{PATH_CHOICE}\{FOURTH_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=6, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # fifth_date
    scr_file = rf'{PATH_CHOICE}\{FIFTH_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=7, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # sixth_date
    scr_file = rf'{PATH_CHOICE}\{SIXTH_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=8, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # seventh_date
    scr_file = rf'{PATH_CHOICE}\{SEVENTH_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=9, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)
createDNsecond()

def createEDS():

    scr_filename = 'EDS-SMART.xlsx'
    dest_filename = 'ЕДС-Смарт Энерджи.xlsx'

    # First date
    scr_file = rf'{PATH_CHOICE}\{FIRST_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=3, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # Second date
    scr_file = rf'{PATH_CHOICE}\{SECOND_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=4, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # third_date
    scr_file = rf'{PATH_CHOICE}\{THIRD_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=5, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # fourth_date
    scr_file = rf'{PATH_CHOICE}\{FOURTH_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=6, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # fifth_date
    scr_file = rf'{PATH_CHOICE}\{FIFTH_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=7, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # sixth_date
    scr_file = rf'{PATH_CHOICE}\{SIXTH_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=8, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # seventh_date
    scr_file = rf'{PATH_CHOICE}\{SEVENTH_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=9, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)
createEDS()

def createLEfirst():

    scr_filename = 'Limanskaya_1.xlsx'
    dest_filename = 'Лиманская Энерджи 1.xlsx'

    # First date
    scr_file = rf'{PATH_CHOICE}\{FIRST_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=3, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # Second date
    scr_file = rf'{PATH_CHOICE}\{SECOND_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=4, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # third_date
    scr_file = rf'{PATH_CHOICE}\{THIRD_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=5, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # fourth_date
    scr_file = rf'{PATH_CHOICE}\{FOURTH_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=6, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # fifth_date
    scr_file = rf'{PATH_CHOICE}\{FIFTH_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=7, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # sixth_date
    scr_file = rf'{PATH_CHOICE}\{SIXTH_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=8, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # seventh_date
    scr_file = rf'{PATH_CHOICE}\{SEVENTH_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=9, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)
createLEfirst()

def createLEsecond():

    scr_filename = 'Limanskaya_2.xlsx'
    dest_filename = 'Лиманская Энерджи 2.xlsx'

    # First date
    scr_file = rf'{PATH_CHOICE}\{FIRST_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=3, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # Second date
    scr_file = rf'{PATH_CHOICE}\{SECOND_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=4, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # third_date
    scr_file = rf'{PATH_CHOICE}\{THIRD_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=5, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # fourth_date
    scr_file = rf'{PATH_CHOICE}\{FOURTH_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=6, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # fifth_date
    scr_file = rf'{PATH_CHOICE}\{FIFTH_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=7, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # sixth_date
    scr_file = rf'{PATH_CHOICE}\{SIXTH_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=8, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # seventh_date
    scr_file = rf'{PATH_CHOICE}\{SEVENTH_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=9, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)
createLEsecond()

def createNeptunSolar():

    scr_filename = 'Neptun_Solar.xlsx'
    dest_filename = 'Нептун Солар.xlsx'

    # First date
    scr_file = rf'{PATH_CHOICE}\{FIRST_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=3, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # Second date
    scr_file = rf'{PATH_CHOICE}\{SECOND_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=4, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # third_date
    scr_file = rf'{PATH_CHOICE}\{THIRD_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=5, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # fourth_date
    scr_file = rf'{PATH_CHOICE}\{FOURTH_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=6, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # fifth_date
    scr_file = rf'{PATH_CHOICE}\{FIFTH_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=7, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # sixth_date
    scr_file = rf'{PATH_CHOICE}\{SIXTH_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=8, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # seventh_date
    scr_file = rf'{PATH_CHOICE}\{SEVENTH_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=9, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)
createNeptunSolar()

def createPlutonSolar():

    scr_filename = 'Pluton_Solar.xlsx'
    dest_filename = 'Плутон Солар.xlsx'

    # First date
    scr_file = rf'{PATH_CHOICE}\{FIRST_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=3, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # Second date
    scr_file = rf'{PATH_CHOICE}\{SECOND_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=4, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # third_date
    scr_file = rf'{PATH_CHOICE}\{THIRD_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=5, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # fourth_date
    scr_file = rf'{PATH_CHOICE}\{FOURTH_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=6, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # fifth_date
    scr_file = rf'{PATH_CHOICE}\{FIFTH_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=7, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # sixth_date
    scr_file = rf'{PATH_CHOICE}\{SIXTH_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=8, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # seventh_date
    scr_file = rf'{PATH_CHOICE}\{SEVENTH_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=9, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)
createPlutonSolar()

def createPRfirst():

    scr_filename = 'Priozernoe_1.xlsx'
    dest_filename = 'Приозерное 1.xlsx'

    # First date
    scr_file = rf'{PATH_CHOICE}\{FIRST_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=3, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # Second date
    scr_file = rf'{PATH_CHOICE}\{SECOND_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=4, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # third_date
    scr_file = rf'{PATH_CHOICE}\{THIRD_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=5, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # fourth_date
    scr_file = rf'{PATH_CHOICE}\{FOURTH_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=6, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # fifth_date
    scr_file = rf'{PATH_CHOICE}\{FIFTH_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=7, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # sixth_date
    scr_file = rf'{PATH_CHOICE}\{SIXTH_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=8, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # seventh_date
    scr_file = rf'{PATH_CHOICE}\{SEVENTH_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=9, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)
createPRfirst()

def createPRsecond():

    scr_filename = 'Priozernoe_2.xlsx'
    dest_filename = 'Приозерное 2.xlsx'

    # First date
    scr_file = rf'{PATH_CHOICE}\{FIRST_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=3, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # Second date
    scr_file = rf'{PATH_CHOICE}\{SECOND_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=4, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # third_date
    scr_file = rf'{PATH_CHOICE}\{THIRD_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=5, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # fourth_date
    scr_file = rf'{PATH_CHOICE}\{FOURTH_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=6, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # fifth_date
    scr_file = rf'{PATH_CHOICE}\{FIFTH_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=7, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # sixth_date
    scr_file = rf'{PATH_CHOICE}\{SIXTH_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=8, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # seventh_date
    scr_file = rf'{PATH_CHOICE}\{SEVENTH_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=9, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)
createPRsecond()

def createFP():

    scr_filename = 'Franko_Pivi.xlsx'
    dest_filename = 'Франко Пиви.xlsx'

    # First date
    scr_file = rf'{PATH_CHOICE}\{FIRST_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=3, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # Second date
    scr_file = rf'{PATH_CHOICE}\{SECOND_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=4, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # third_date
    scr_file = rf'{PATH_CHOICE}\{THIRD_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=5, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # fourth_date
    scr_file = rf'{PATH_CHOICE}\{FOURTH_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=6, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # fifth_date
    scr_file = rf'{PATH_CHOICE}\{FIFTH_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=7, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # sixth_date
    scr_file = rf'{PATH_CHOICE}\{SIXTH_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=8, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # seventh_date
    scr_file = rf'{PATH_CHOICE}\{SEVENTH_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=9, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)
createFP()

def createFS():

    scr_filename = 'Franko_Solar.xlsx'
    dest_filename = 'Франко Солар.xlsx'

    # First date
    scr_file = rf'{PATH_CHOICE}\{FIRST_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=3, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # Second date
    scr_file = rf'{PATH_CHOICE}\{SECOND_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=4, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # third_date
    scr_file = rf'{PATH_CHOICE}\{THIRD_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=5, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # fourth_date
    scr_file = rf'{PATH_CHOICE}\{FOURTH_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=6, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # fifth_date
    scr_file = rf'{PATH_CHOICE}\{FIFTH_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=7, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # sixth_date
    scr_file = rf'{PATH_CHOICE}\{SIXTH_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=8, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # seventh_date
    scr_file = rf'{PATH_CHOICE}\{SEVENTH_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=9, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)
createFS()

def createEvda():

    scr_filename = 'Evda_Energo.xlsx'
    dest_filename = 'Эвда Энерго.xlsx'

    # First date
    scr_file = rf'{PATH_CHOICE}\{FIRST_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=3, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # Second date
    scr_file = rf'{PATH_CHOICE}\{SECOND_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=4, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # third_date
    scr_file = rf'{PATH_CHOICE}\{THIRD_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=5, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # fourth_date
    scr_file = rf'{PATH_CHOICE}\{FOURTH_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=6, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # fifth_date
    scr_file = rf'{PATH_CHOICE}\{FIFTH_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=7, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # sixth_date
    scr_file = rf'{PATH_CHOICE}\{SIXTH_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=8, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)

    # seventh_date
    scr_file = rf'{PATH_CHOICE}\{SEVENTH_DATE}\{scr_filename}'
    book_scr = load_workbook(filename=scr_file)
    sheet_book_scr = book_scr.active

    dest_file = rf'{DEST_DIR}\{dest_filename}'
    book_dest = load_workbook(filename=dest_file)
    sheet_book_dest = book_dest.active

    for row in range(3, 28):
        cell_book_scr = sheet_book_scr.cell(row=row, column=3)
        sheet_book_dest.cell(row=row + 2, column=9, value=cell_book_scr.value)

    book_dest.save(filename=dest_file)
createEvda()

easygui.msgbox('Звіт успішно створено!')