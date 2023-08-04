# -*- coding: utf-8 -*-
import os
import shutil
from datetime import datetime, timedelta
import easygui
from openpyxl import load_workbook, Workbook

TEMPLATES_DIR = r"C:\Users\ihoraryku\Downloads\weekly_report\templates"
PATH_CHOICE = None  # Move PATH_CHOICE to the global scope


def get_next_seven_dates():
    today = datetime.utcnow().date()
    date_range = [today + timedelta(days=i) for i in range(1, 8)]
    date_strings = [date.strftime("%d.%m.%Y") for date in date_range]
    print(f"Date_strings: {date_strings}")
    return date_strings


def copy_templates(src_dir_templates, destination):
    try:
        if os.path.exists(destination):
            shutil.rmtree(destination)
        shutil.copytree(src_dir_templates, destination)
    except Exception as err:
        print(f"Error copying templates: {err}")


def process_report(source_file, destination_file, column_index, date_strings, dest_dir):
    dest_file = os.path.join(dest_dir, destination_file)
    print(f"dest_file: {dest_file}")

    try:
        if os.path.exists(dest_file):
            book_dest = load_workbook(filename=dest_file)
            sheet_dest = book_dest.active
        else:
            book_dest = Workbook()
            sheet_dest = book_dest.active

        for i, date_str in enumerate(date_strings, start=1):
            src_file = os.path.join(PATH_CHOICE, date_str, source_file)
            print(f"source_file: {src_file}")
            book_src = load_workbook(filename=src_file)
            sheet_src = book_src.active

            for row in range(3, 28):
                cell_src = sheet_src.cell(row=row, column=3)
                sheet_dest.cell(row=row + 2, column=column_index + i, value=cell_src.value)

        book_dest.save(filename=dest_file)

    except Exception as err:
        print(f"Error processing report: {err}")


def main():
    global PATH_CHOICE
    easygui.msgbox(
        'Вас вітає програма для автоматизованого створення щонедільного звіту для НЕК. \nОберіть необхідну теку (місяць) де зберігаються файли НЕК'
    )
    PATH_CHOICE = easygui.diropenbox()

    first_date = (datetime.utcnow().date() + timedelta(1)).strftime("%d.%m.%Y")
    seventh_date = (datetime.utcnow().date() + timedelta(7)).strftime("%d.%m.%Y")
    dest_dir = os.path.join(PATH_CHOICE, f'Недельный на {first_date}-{seventh_date}')
    print(f'path_choice: {PATH_CHOICE}')
    print(f'dest_dir: {dest_dir}')

    date_strings = get_next_seven_dates()

    copy_templates(TEMPLATES_DIR, dest_dir)

    reports = [
        ("Bolgrad_Solar.xlsx", "Болград Солар.xlsx"),
        ("Voshod_Solar.xlsx", "Восход Солар.xlsx"),
        ("Dunaiskaya_SES1.xlsx", "Дунайская СЭС-1.xlsx"),
        ("Dunaiskaya_SES2.xlsx", "Дунайская СЭС-2.xlsx"),
        ("EDS-SMART.xlsx", "ЕДС-Смарт Энерджи.xlsx"),
        ("Limanskaya_1.xlsx", "Лиманская Энерджи 1.xlsx"),
        ("Limanskaya_2.xlsx", "Лиманская Энерджи 2.xlsx"),
        ("Neptun_Solar.xlsx", "Нептун Солар.xlsx"),
        ("Pluton_Solar.xlsx", "Плутон Солар.xlsx"),
        ("Priozernoe_1.xlsx", "Приозерное 1.xlsx"),
        ("Priozernoe_2.xlsx", "Приозерное 2.xlsx"),
        ("Franko_Pivi.xlsx", "Франко Пиви.xlsx"),
        ("Franko_Solar.xlsx", "Франко Солар.xlsx"),
        ("Evda_Energo.xlsx", "Эвда Энерго.xlsx")
    ]

    for source_file, destination_file in reports:
        process_report(source_file, destination_file, 2, date_strings, dest_dir)

    easygui.msgbox('Звіт успішно створено!')


if __name__ == "__main__":
    main()